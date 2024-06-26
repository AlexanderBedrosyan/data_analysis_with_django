import json

from django.shortcuts import render
from django.views import View
from openpyxl import load_workbook
from io import BytesIO
from .objects_of_companies.companies_db import AllCompanies, SingleCompany
from django.http import HttpResponse

# Create your views here.


class Home(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'home.html')

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('fileInput')
        if not file:
            return HttpResponse("You have to choose file!", status=400)

        file_content = BytesIO(file.read())
        workbook = load_workbook(file_content)
        sheet_names = workbook.sheetnames

        all_companies = AllCompanies()

        for company in sheet_names:
            selected_sheet_name = company
            sheet = workbook[selected_sheet_name]

            new_company = SingleCompany(company)

            for row in sheet.iter_rows(values_only=True):
                current_name = row[1]
                new_company.check_if_company_exist(current_name)

                invoice_number = row[0]
                price = float(row[2])
                new_company.add_invoices(invoice_number, price, current_name)

            all_companies.add_new_company_object(new_company)

        result = all_companies.difference_checker()
        print(result)

        return render(request, 'home.html', {"result": json.dumps(result)})